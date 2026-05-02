# -*- coding: utf-8 -*-
"""
舟谱系统自提订单导入模板生成器

从下单表 + 价格表生成舟谱系统可导入的自提订单Excel文件。

用法:
  python generate_order_import.py --order <下单表xlsx> --price <价格表xlsx> --arrival <到货日期YYYYMMDD> [--sheet <工作表名关键词>] [--start-seq <起始序号>] [--output <输出路径>] [--extra-prices <额外价格JSON>]

参数:
  --order        下单表Excel文件路径（必填）
  --price        价格表Excel文件路径（必填）
  --arrival      到货日期，格式YYYYMMDD（必填）
  --sheet        下单表工作表名关键词，用于筛选工作表（可选，默认包含到货日期数字的工作表）
  --start-seq    源单据号起始序号，默认21（避开舟谱系统已有单号01~20）
  --output       输出文件路径（可选，默认: 自提订单导入模板_<到货日期>.xlsx）
  --extra-prices 额外价格JSON字符串，格式: {"条码":{"美联价格":10.47,"永辉价格":12.9}}（可选）
  --config       客户配置JSON文件路径（可选，默认使用内置配置）

示例:
  python generate_order_import.py --order 下单表.xlsx --price 价格表.xlsx --arrival 20260408
  python generate_order_import.py --order 下单表.xlsx --price 价格表.xlsx --arrival 20260408 --start-seq 1 --extra-prices '{"6970618572449":{"美联价格":10.47}}'
"""

import pandas as pd
from openpyxl import Workbook
import openpyxl
from openpyxl.styles import Font, Alignment
import argparse
import json
import sys
import os

# 强制UTF-8输出
if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')
if sys.stderr.encoding and sys.stderr.encoding.lower() != 'utf-8':
    sys.stderr.reconfigure(encoding='utf-8')

# ============================================================
# 内置客户配置（可被 --config 覆盖）
# ============================================================
DEFAULT_CONFIG = {
    # 分销商列表（下单表中的列名）
    "distributors": ["唐成", "黄家伟", "易胜琳", "胡魁魁", "朱青峰", "谢总"],
    # 门店列表（下单表中的列名）
    "stores": ["吾悦", "东津", "民发", "沃尔玛", "檀溪美联"],
    # 客户名称映射：下单表列名 → 舟谱系统客户名称
    "customer_mapping": {
        "易胜琳": "易胜玲",
        "谢总": "宜城谢总",
        "吾悦": "永辉吾悦店",
        "东津": "永辉东津店",
        "民发": "永辉民发店",
        "檀溪美联": "美联檀溪店"
    },
    # 客户类型 → 价格列名映射规则
    # 分销商用"分销价格"，永辉门店用"永辉价格"，等
    "price_rules": {
        "唐成": "分销价格",
        "黄家伟": "分销价格",
        "易胜琳": "分销价格",
        "胡魁魁": "分销价格",
        "朱青峰": "分销价格",
        "谢总": "分销价格",
        "吾悦": "永辉价格",
        "东津": "永辉价格",
        "民发": "永辉价格",
        "沃尔玛": "沃尔玛价格",
        "檀溪美联": "美联价格"
    },
    # 固定字段
    "业务员": "张俊峰",
    "部门": "湖北福宝商贸有限公司",
    "仓库": "总仓"
}


def load_config(config_path=None):
    """加载客户配置，合并默认值"""
    config = json.loads(json.dumps(DEFAULT_CONFIG))  # deep copy
    if config_path and os.path.exists(config_path):
        with open(config_path, 'r', encoding='utf-8') as f:
            user_config = json.load(f)
        config.update(user_config)
    return config


def _is_nan(val):
    """安全判断NaN，兼容pandas和openpyxl"""
    if val is None:
        return True
    if isinstance(val, float):
        return val != val  # NaN check
    return False


def load_price_table(price_file):
    """读取价格表，返回 {条码: {单位, 简称, 各价格}} 字典"""
    import openpyxl
    wb = openpyxl.load_workbook(price_file, data_only=True)
    ws = wb['数据源']
    barcode_info = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        # 表头顺序: 代码, 条码, 永辉代码, 简称, 规格, 单位, 分销价格, 永辉价格, 沃尔玛价格, 美联价格
        barcode = row[1]  # 条码列
        if _is_nan(barcode):
            continue
        barcode = str(int(barcode)) if isinstance(barcode, float) else str(barcode).strip()
        if barcode not in barcode_info:
            barcode_info[barcode] = {
                '单位': row[5] if not _is_nan(row[5]) else None,
                '简称': row[3] if not _is_nan(row[3]) else None,
                '分销价格': _safe_round(row[6]),
                '永辉价格': _safe_round(row[7]),
                '沃尔玛价格': _safe_round(row[8]),
                '美联价格': _safe_round(row[9]),
            }
        else:
            # 补充空值
            for col_idx, col_key in enumerate(['分销价格', '永辉价格', '沃尔玛价格', '美联价格']):
                if barcode_info[barcode][col_key] is None:
                    val = _safe_round(row[6 + col_idx])
                    if val is not None:
                        barcode_info[barcode][col_key] = val
    wb.close()
    return barcode_info


def _safe_round(val):
    """安全四舍五入，处理NaN"""
    if val is None:
        return None
    if isinstance(val, float) and val != val:  # NaN
        return None
    try:
        return round(float(val), 2)
    except (ValueError, TypeError):
        return None


def apply_extra_prices(barcode_info, extra_prices):
    """应用额外的价格补充"""
    if not extra_prices:
        return barcode_info
    for barcode, data in extra_prices.items():
        barcode = str(barcode).strip()
        if barcode not in barcode_info:
            barcode_info[barcode] = {'单位': data.get('单位', ''), '简称': data.get('简称', ''), '分销价格': None, '永辉价格': None, '沃尔玛价格': None, '美联价格': None}
        for key, val in data.items():
            if key in ('分销价格', '永辉价格', '沃尔玛价格', '美联价格') and val is not None:
                barcode_info[barcode][key] = round(float(val), 2)
            elif key == '单位' and val:
                barcode_info[barcode]['单位'] = val
            elif key == '简称' and val:
                barcode_info[barcode]['简称'] = val
    return barcode_info


def _read_excel_rows(file_path, sheet_name):
    """用openpyxl读取Excel工作表，返回(表头列表, 行数据列表)"""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"未找到工作表: {sheet_name}")
    ws = wb[sheet_name]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            rows.append(dict(zip(headers, row)))
    wb.close()
    return headers, rows


def _list_sheets(file_path):
    """列出Excel所有工作表名"""
    wb = openpyxl.load_workbook(file_path, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def select_sheet(order_file, sheet_keyword=None, arrival_date=None):
    """选择下单表工作表"""
    all_sheets = _list_sheets(order_file)

    if sheet_keyword:
        matched = [s for s in all_sheets if sheet_keyword in s]
        if matched:
            return matched[0]
        print(f"⚠️ 未找到包含'{sheet_keyword}'的工作表，可用工作表: {all_sheets}")

    # 根据到货日期筛选
    if arrival_date:
        # 提取月日数字，如 20260408 → "4月8" 或 "08"
        month = str(int(arrival_date[4:6]))
        day = str(int(arrival_date[6:8]))
        for s in all_sheets:
            if month in s and day in s:
                return s

    # 默认返回最后一个工作表（通常是最新报单）
    if len(all_sheets) == 1:
        return all_sheets[0]

    print(f"⚠️ 多个工作表: {all_sheets}")
    print(f"   请用 --sheet 参数指定工作表名关键词")
    return None


def process_order_sheet(rows, config, barcode_info):
    """处理下单表，返回 {客户名称: [{条码, 单位, 数量, 单价}]} 字典"""
    # 从第一行获取列名
    if not rows:
        return {}, [], []
    all_cols = list(rows[0].keys()) if isinstance(rows[0], dict) else []
    customer_cols = [c for c in all_cols
                     if c in config['distributors'] + config['stores']]

    if not customer_cols:
        print(f"⚠️ 下单表中未找到匹配的客户列。现有列: {all_cols}")
        return {}, [], []

    customer_orders = {}
    missing_price = []
    missing_barcode = []

    for row in rows:
        barcode = row.get('条码')
        if _is_nan(barcode):
            continue
        barcode_str = str(int(barcode)) if isinstance(barcode, float) else str(barcode).strip()
        info = barcode_info.get(barcode_str)

        for col in customer_cols:
            qty = row.get(col)
            try:
                qty_num = float(str(qty).strip())
            except (ValueError, TypeError):
                continue
            if qty_num <= 0:
                continue

            # 映射客户名称
            customer_name = config['customer_mapping'].get(col, col)

            # 获取单位
            unit = ''
            if info:
                unit_val = info.get('单位')
                unit = str(unit_val) if unit_val and not _is_nan(unit_val) else ''
            else:
                if barcode_str not in missing_barcode:
                    missing_barcode.append(barcode_str)

            # 获取价格
            price_col = config['price_rules'].get(col, '分销价格')
            price = None
            if info:
                price = info.get(price_col)
                if price is None:
                    missing_price.append(f"{barcode_str} {info.get('简称', '')} → {customer_name}无{price_col}")

            if customer_name not in customer_orders:
                customer_orders[customer_name] = []
            customer_orders[customer_name].append({
                '条码': int(barcode_str),
                '单位': str(unit) if unit else '',
                '数量': int(qty_num),
                '单价': float(price) if price else None
            })

    return customer_orders, missing_price, missing_barcode


def generate_excel(customer_orders, config, arrival_date, start_seq=21):
    """生成舟谱导入格式的Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = '自提订单'

    headers = [
        '*源单据号', '客户编号', '客户名称', '*业务员', '部门', '*仓库', '单据日期',
        '整单备注', '制单人', '商品编号', '商品货号', '商品名称', '商品条码',
        '*单位', '*数量', '*单价(折后价)', '明细备注', '业务属性', '标签'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row_idx = 2
    order_counter = start_seq

    for customer_name in sorted(customer_orders.keys()):
        # 源单据号格式: ZT + 日期 + 序号（如 ZT2026040801）
        order_no = f"ZT{arrival_date}{str(order_counter).zfill(2)}"
        order_counter += 1
        for item in customer_orders[customer_name]:
            ws.cell(row=row_idx, column=1, value=order_no)      # *源单据号
            ws.cell(row=row_idx, column=3, value=customer_name)   # 客户名称
            ws.cell(row=row_idx, column=4, value=config['业务员'])  # *业务员
            ws.cell(row=row_idx, column=5, value=config['部门'])    # 部门
            ws.cell(row=row_idx, column=6, value=config['仓库'])    # *仓库
            ws.cell(row=row_idx, column=13, value=item['条码'])    # 商品条码
            ws.cell(row=row_idx, column=14, value=item['单位'])    # *单位
            ws.cell(row=row_idx, column=15, value=item['数量'])    # *数量
            ws.cell(row=row_idx, column=16, value=item['单价'])    # *单价(折后价)
            row_idx += 1

    return wb, row_idx - 2


def verify_required_fields(output_file):
    """验证所有必填字段无空值"""
    _, rows = _read_excel_rows(output_file, '自提订单')
    required = ['*源单据号', '*业务员', '*仓库', '*单位', '*数量', '*单价(折后价)']
    all_ok = True
    issues = []
    for col in required:
        null_rows = [r for r in rows if _is_nan(r.get(col))]
        if null_rows:
            all_ok = False
            for r in null_rows:
                issues.append(f"  {col}为空: 客户={r.get('客户名称','')}, 条码={r.get('商品条码','')}")
    return all_ok, issues


def main():
    parser = argparse.ArgumentParser(description='舟谱系统自提订单导入模板生成器')
    parser.add_argument('--order', required=True, help='下单表Excel文件路径')
    parser.add_argument('--price', required=True, help='价格表Excel文件路径')
    parser.add_argument('--arrival', required=True, help='到货日期YYYYMMDD')
    parser.add_argument('--sheet', default=None, help='工作表名关键词')
    parser.add_argument('--start-seq', type=int, default=21, help='源单据号起始序号（默认21）')
    parser.add_argument('--output', default=None, help='输出文件路径')
    parser.add_argument('--extra-prices', default=None, help='额外价格JSON字符串')
    parser.add_argument('--extra-prices-file', default=None, help='额外价格JSON文件路径')
    parser.add_argument('--config', default=None, help='客户配置JSON文件路径')
    args = parser.parse_args()

    print(f"📅 到货日期: {args.arrival}")
    print(f"📄 下单表: {args.order}")
    print(f"💰 价格表: {args.price}")

    # 1. 加载配置
    config = load_config(args.config)

    # 2. 读取价格表
    print("\n📖 读取价格表...")
    barcode_info = load_price_table(args.price)
    print(f"   共 {len(barcode_info)} 个条码")

    # 3. 应用额外价格
    extra = {}
    if args.extra_prices:
        extra.update(json.loads(args.extra_prices))
    if args.extra_prices_file and os.path.exists(args.extra_prices_file):
        with open(args.extra_prices_file, 'r', encoding='utf-8') as f:
            extra.update(json.load(f))
    if extra:
        barcode_info = apply_extra_prices(barcode_info, extra)
        print(f"   补充 {len(extra)} 个条码的价格")

    # 4. 选择工作表
    print("\n📖 读取下单表...")
    sheet_name = select_sheet(args.order, args.sheet, args.arrival)
    if sheet_name is None:
        sys.exit(1)
    print(f"   工作表: {sheet_name}")

    _, rows = _read_excel_rows(args.order, sheet_name)
    print(f"   共 {len(rows)} 行")

    # 5. 处理订单数据
    customer_orders, missing_price, missing_barcode = process_order_sheet(rows, config, barcode_info)

    if missing_barcode:
        print(f"\n⚠️ 以下条码在价格表中不存在:")
        for b in set(missing_barcode):
            print(f"   {b}")

    if missing_price:
        print(f"\n⚠️ 以下条码缺少对应客户类型的价格:")
        for m in missing_price:
            print(f"   {m}")
        print("\n请用 --extra-prices 参数补充缺失价格，格式:")
        print('  --extra-prices \'{"条码":{"价格列名":价格值}}\'')
        print("  价格列名: 分销价格、永辉价格、沃尔玛价格、美联价格")

    if not customer_orders:
        print("\n❌ 未生成任何订单记录，请检查下单表和客户列名是否匹配")
        sys.exit(1)

    # 6. 生成Excel
    total_records = sum(len(items) for items in customer_orders.values())
    print(f"\n📦 生成订单: {len(customer_orders)} 个客户, {total_records} 条记录")

    wb, _ = generate_excel(customer_orders, config, args.arrival, args.start_seq)

    output = args.output or f"自提订单导入模板_{args.arrival}.xlsx"
    wb.save(output)
    print(f"💾 已保存: {output}")

    # 7. 验证必填字段
    all_ok, issues = verify_required_fields(output)
    if all_ok:
        print("✅ 所有必填字段均无空值")
    else:
        print("❌ 必填字段有空值:")
        for issue in issues:
            print(issue)

    # 8. 订单统计
    print(f"\n=== 订单统计 ===")
    for name in sorted(customer_orders.keys()):
        items = customer_orders[name]
        print(f"  {name}: {len(items)}条")
        if any(item['单价'] is None for item in items):
            null_items = [item for item in items if item['单价'] is None]
            print(f"    ⚠️ 其中 {len(null_items)} 条缺少单价")


if __name__ == '__main__':
    main()
