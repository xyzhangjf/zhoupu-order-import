# -*- coding: utf-8 -*-
"""
舟谱系统调拨订单导入模板生成器

从下单表 + 价格表生成舟谱系统可导入的调拨订单Excel文件。

用法:
  python generate_transfer_order.py --order <下单表xlsx> --price <价格表xlsx> --arrival <到货日期YYYYMMDD> [--sheet <工作表名关键词>] [--start-seq <起始序号>] [--output <输出路径>]

参数:
  --order        下单表Excel文件路径（必填）
  --price        价格表Excel文件路径（必填，用于获取条码对应单位）
  --arrival      到货日期，格式YYYYMMDD（必填）
  --sheet        下单表工作表名关键词（可选，默认自动匹配到货日期）
  --start-seq    源单据号起始序号，默认1
  --output       输出文件路径（可选，默认: 调拨订单导入模板_<到货日期>.xlsx）
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import argparse
import sys

if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')
if sys.stderr.encoding and sys.stderr.encoding.lower() != 'utf-8':
    sys.stderr.reconfigure(encoding='utf-8')

# 调拨业务员列表（下单表中的列名）
TRANSFER_STAFF = ["程欢欢", "刘善涛", "毛辉", "周运潘", "田顺达", "王琴", "刘正宝"]

# 固定值
TRANSFER_OUT_WAREHOUSE = "总仓"


def _is_nan(val):
    """安全判断NaN，兼容openpyxl"""
    if val is None:
        return True
    if isinstance(val, float) and val != val:  # NaN
        return True
    return False


def _read_excel_rows(file_path, sheet_name):
    """用openpyxl读取Excel工作表，返回(表头列表, 行数据列表)"""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"未找到工作表: {sheet_name}")
    ws = wb[sheet_name]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            rows.append(dict(zip(headers, row)))
    wb.close()
    return headers, rows


def _list_sheets(file_path):
    """列出Excel所有工作表名"""
    wb = openpyxl.load_workbook(file_path, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def load_price_table(price_file):
    """读取价格表，返回 {条码: {单位, 简称}} 字典（调拨不需要价格，只需要单位和条码信息）"""
    wb = openpyxl.load_workbook(price_file, data_only=True)
    ws = wb['数据源']
    barcode_info = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        barcode = row[1]  # 条码列
        if _is_nan(barcode):
            continue
        barcode = str(int(barcode)) if isinstance(barcode, float) else str(barcode).strip()
        if barcode not in barcode_info:
            barcode_info[barcode] = {
                '单位': row[5] if not _is_nan(row[5]) else None,  # 单位
                '简称': row[3] if not _is_nan(row[3]) else None,  # 简称
            }
    wb.close()
    return barcode_info


def select_sheet(order_file, sheet_keyword=None, arrival_date=None):
    """选择下单表工作表"""
    all_sheets = _list_sheets(order_file)

    if sheet_keyword:
        matched = [s for s in all_sheets if sheet_keyword in s]
        if matched:
            return matched[0]
        print(f"⚠️ 未找到包含'{sheet_keyword}'的工作表，可用工作表: {all_sheets}")

    if arrival_date:
        month = str(int(arrival_date[4:6]))
        day = str(int(arrival_date[6:8]))
        for s in all_sheets:
            if month in s and day in s:
                return s

    if len(all_sheets) == 1:
        return all_sheets[0]

    print(f"⚠️ 多个工作表: {all_sheets}")
    print(f"   请用 --sheet 参数指定工作表名关键词")
    return None


def process_order_sheet(rows, barcode_info):
    """处理下单表，提取调拨业务员的订单数据"""
    if not rows:
        return {}, []
    all_cols = list(rows[0].keys())
    staff_cols = [c for c in all_cols if c in TRANSFER_STAFF]

    if not staff_cols:
        print(f"⚠️ 下单表中未找到调拨业务员列。现有列: {all_cols}")
        print(f"   期望找到: {TRANSFER_STAFF}")
        return {}, []

    staff_orders = {}
    missing_barcode = []

    for row in rows:
        barcode = row.get('条码')
        if _is_nan(barcode):
            continue
        barcode_str = str(int(barcode)) if isinstance(barcode, float) else str(barcode).strip()
        info = barcode_info.get(barcode_str)

        for col in staff_cols:
            qty = row.get(col)
            # 转换为数字
            try:
                qty_num = int(float(str(qty).strip()))
            except (ValueError, TypeError):
                continue
            if qty_num <= 0:
                continue

            # 调入仓 = 业务员名 + "仓"
            warehouse_in = col + "仓"

            # 获取单位
            unit = ''
            if info:
                unit_val = info.get('单位')
                unit = str(unit_val) if unit_val and not _is_nan(unit_val) else ''
            else:
                if barcode_str not in missing_barcode:
                    missing_barcode.append(barcode_str)

            if warehouse_in not in staff_orders:
                staff_orders[warehouse_in] = []
            staff_orders[warehouse_in].append({
                '条码': int(barcode_str),
                '单位': str(unit) if unit else '',
                '数量': qty_num,
            })

    return staff_orders, missing_barcode


def generate_excel(staff_orders, arrival_date, start_seq=1):
    """生成舟谱调拨订单导入格式的Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = '调拨订单'

    headers = [
        '*源单据号', '*调出仓', '*调入仓', '*单位', '整单备注',
        '商品编号', '商品名称', '商品条码', '期望生产日期',
        '*订单数量', '调拨参考价', '明细备注'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row_idx = 2
    seq = start_seq

    for warehouse_in in sorted(staff_orders.keys()):
        order_no = f"DB{arrival_date}{str(seq).zfill(2)}"
        seq += 1
        for item in staff_orders[warehouse_in]:
            ws.cell(row=row_idx, column=1, value=order_no)            # *源单据号
            ws.cell(row=row_idx, column=2, value=TRANSFER_OUT_WAREHOUSE)  # *调出仓
            ws.cell(row=row_idx, column=3, value=warehouse_in)          # *调入仓
            ws.cell(row=row_idx, column=4, value=item['单位'])          # *单位
            ws.cell(row=row_idx, column=8, value=item['条码'])          # 商品条码
            ws.cell(row=row_idx, column=10, value=item['数量'])         # *订单数量
            row_idx += 1

    return wb, row_idx - 2


def verify_required_fields(output_file):
    """验证所有必填字段无空值"""
    _, rows = _read_excel_rows(output_file, '调拨订单')
    required = ['*源单据号', '*调出仓', '*调入仓', '*单位', '*订单数量', '商品条码']
    all_ok = True
    issues = []
    for col in required:
        null_rows = [r for r in rows if _is_nan(r.get(col))]
        if null_rows:
            all_ok = False
            for r in null_rows:
                issues.append(f"  {col}为空: 调入仓={r.get('*调入仓','')}, 条码={r.get('商品条码','')}")
    return all_ok, issues


def main():
    parser = argparse.ArgumentParser(description='舟谱系统调拨订单导入模板生成器')
    parser.add_argument('--order', required=True, help='下单表Excel文件路径')
    parser.add_argument('--price', required=True, help='价格表Excel文件路径')
    parser.add_argument('--arrival', required=True, help='到货日期YYYYMMDD')
    parser.add_argument('--sheet', default=None, help='工作表名关键词')
    parser.add_argument('--start-seq', type=int, default=1, help='源单据号起始序号（默认1）')
    parser.add_argument('--output', default=None, help='输出文件路径')
    args = parser.parse_args()

    print(f"📅 到货日期: {args.arrival}")
    print(f"📄 下单表: {args.order}")
    print(f"💰 价格表: {args.price}")

    # 1. 读取价格表（获取条码→单位映射）
    print("\n📖 读取价格表...")
    barcode_info = load_price_table(args.price)
    print(f"   共 {len(barcode_info)} 个条码")

    # 2. 选择工作表
    print("\n📖 读取下单表...")
    sheet_name = select_sheet(args.order, args.sheet, args.arrival)
    if sheet_name is None:
        sys.exit(1)
    print(f"   工作表: {sheet_name}")

    _, rows = _read_excel_rows(args.order, sheet_name)
    print(f"   共 {len(rows)} 行")

    # 3. 处理订单数据
    staff_orders, missing_barcode = process_order_sheet(rows, barcode_info)

    if missing_barcode:
        print(f"\n⚠️ 以下条码在价格表中不存在:")
        for b in set(missing_barcode):
            print(f"   {b}")

    if not staff_orders:
        print("\n❌ 未生成任何调拨订单记录，请检查下单表中是否有调拨业务员的下单数据")
        sys.exit(1)

    # 4. 生成Excel
    total_records = sum(len(items) for items in staff_orders.values())
    print(f"\n📦 生成调拨订单: {len(staff_orders)} 个仓库, {total_records} 条记录")

    wb, _ = generate_excel(staff_orders, args.arrival, args.start_seq)

    output = args.output or f"调拨订单导入模板_{args.arrival}.xlsx"
    wb.save(output)
    print(f"💾 已保存: {output}")

    # 5. 验证必填字段
    all_ok, issues = verify_required_fields(output)
    if all_ok:
        print("✅ 所有必填字段均无空值")
    else:
        print("❌ 必填字段有空值:")
        for issue in issues:
            print(issue)

    # 6. 订单统计
    print(f"\n=== 订单统计 ===")
    for name in sorted(staff_orders.keys()):
        items = staff_orders[name]
        print(f"  {name}: {len(items)}条")


if __name__ == '__main__':
    main()
